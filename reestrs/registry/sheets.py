# -*- coding: utf-8 -*-

from string import ascii_uppercase

from peewee import fn
from playhouse.shortcuts import case
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THICK, BORDER_THIN
from openpyxl.utils import get_column_letter

from utils import get_period_range, previous_month_day, get_day_range, get_dynamic_range, format_date_range, format_date


# The number format of circulation
CIRC_NUMBER_FORMAT = '#,##0'

DYNAMIC_NUMBER_FORMAT = '0%'


def calculate_dynamic(current, prev):
    formula = "=IF({a}=0,-{b},IF({b}=0,{a},({a}-{b})/{b}))"
    return formula.format(a=current.coordinate, b=prev.coordinate)


def calculations_by_dates(dt, model, calculated_field):
    """ Производится подсчет кол-ва и суммы оборотов для разных дат
    param: model - Модель Invoice или Withdraw
    param: calculated_field - Поле модели, по которому будет производиться расчет оборота
        (например Invoice.shop_amount)
    """
    current_day = get_day_range(dt)
    current_period = get_period_range(dt)
    prev_month_day = get_day_range(previous_month_day(dt))
    prev_month_period = get_period_range(previous_month_day(dt))
    return (
        # Подсчет оборота за текущий день
        fn.SUM(case(None, ((model.processed.between(*current_day),
                            calculated_field),), 0)).alias('curr_day_circ'),

        # Количество инвойсов за текущий день
        fn.COUNT(case(None, ((model.processed.between(*current_day),
                             model.id),), None)).alias('curr_day_count'),

        # Оборот за текущий период
        fn.SUM(case(None, ((model.processed.between(*current_period),
                           calculated_field),), 0)).alias('curr_period_circ'),

        # Количество инвойсов за текущий период
        fn.COUNT(case(None, ((model.processed.between(*current_period),
                              model.id),), None)).alias('curr_period_count'),

        # Оборот за текущий день в прошлом месяце
        fn.SUM(case(None, ((model.processed.between(*prev_month_day),
                            calculated_field),), 0)).alias('prev_day_circ'),

        # Количество инвойсов за текущий период в прошлом месяце
        fn.COUNT(case(None, ((model.processed.between(*prev_month_day),
                              model.id),), None)).alias('prev_day_count'),

        # Оборот за период в прошедшем месяце
        fn.SUM(case(None, ((model.processed.between(*prev_month_period),
                            calculated_field),), 0)).alias('prev_period_circ'),

        # Количество инвойсов за период в прошедшем месяце
        fn.COUNT(case(None, ((model.processed.between(*prev_month_period),
                              model.id),), None)).alias('prev_period_count'),
    )


class Sheet(object):
    def __init__(self, wb, title, currency, date):
        self._sheet = wb.create_sheet(title)
        self._currency = currency
        self._date = date
        self.dynamic_date = get_dynamic_range(self._date, previous_month_day(self._date))
        self.create_table()

    def create_table(self):
        self._create_table_header()
        self._fill_table()
        self._create_table_footer()

    def _create_table_header(self):
        raise NotImplementedError()

    def _fill_table(self):
        raise NotImplementedError()

    def _create_table_footer(self):
        sheet = self._sheet
        max_row = sheet.max_row
        last_row = max_row + 2  # Skip two lines after last raw

        sheet['C%s' % last_row] = u'ИТОГО'
        cell_chars = ('D', 'E', 'G', 'H', 'J', 'K', 'L', 'M')
        for char in sorted(cell_chars):
            sheet['{0}{1}'.format(char, last_row)] = u'=SUM({0}3:{1}{2})'.format(char, char, max_row)

        sheet['D%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['E%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['F%s' % last_row] = calculate_dynamic(sheet['D%s' % last_row], sheet['E%s' % last_row])
        sheet['F%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT
        sheet['G%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['H%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['I%s' % last_row] = calculate_dynamic(sheet['G%s' % last_row], sheet['H%s' % last_row])
        sheet['I%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT

        self._apply_border()

    def _format_header(self):
        sheet = self._sheet
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 30
        for letter in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
            sheet.column_dimensions[letter].width = 20

        for cell in sheet.get_cell_collection():
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    def _get_border(self, left_side, right_side, top_side, bottom_side):
        return Border(left=Side(style=left_side),
                      right=Side(style=right_side),
                      top=Side(style=top_side),
                      bottom=Side(style=bottom_side))

    def _apply_border(self):
        sheet = self._sheet
        for row in sheet.rows:
            for cell in row:
                cell.border = self._get_border(BORDER_THIN, BORDER_THIN, BORDER_THIN, BORDER_THIN)

        max_column = get_column_letter(sheet.max_column)
        sum_record = 'A{0}:{1}{2}'.format(sheet.max_row, max_column, sheet.max_row)
        for row in sheet.iter_rows(sum_record):
            for cell in row:
                cell.border = self._get_border(BORDER_THIN, BORDER_THIN, BORDER_THICK, BORDER_THICK)

        self._add_right_vertical_borders()

    def _add_right_vertical_borders(self):
        letters = ("C", "F", "I", "K", "M")
        map(self._add_vertical_border, letters)

    def _add_vertical_border(self, letter):
        sheet = self._sheet
        if letter not in ascii_uppercase:
            raise Exception('Letter %s must be in uppercase' % letter)

        row = "{0}1:{1}{2}".format(letter, letter, sheet.max_row - 1)
        for row in sheet.iter_rows(row):
            for cell in row:
                cell.border = self._get_border(BORDER_THIN, BORDER_THICK, BORDER_THIN, BORDER_THIN)


class ShopSheet(Sheet):
    def _create_table_header(self):
        sheet = self._sheet
        sheet['A1'] = u'ID магазина'
        sheet['B1'] = u'URL магазина'
        sheet['C1'] = u'Название магазина'
        sheet['D1'] = format_date_range(get_period_range(self._date))
        sheet['D2'] = u'Оборот'

        prev_month_period = get_period_range(previous_month_day(self._date))
        sheet['E1'] = format_date_range(prev_month_period)
        sheet['E2'] = u'Оборот'

        sheet['F1'] = self.dynamic_date
        sheet['F2'] = u'Динамика'

        sheet['G1'] = format_date(self._date)
        sheet['G2'] = u'Оборот'

        sheet['H1'] = format_date(previous_month_day(self._date))
        sheet['H2'] = u'Оборот'

        sheet['I1'] = self.dynamic_date
        sheet['I2'] = u'Динамика'

        sheet['J1'] = format_date_range(get_period_range(self._date))
        sheet['J2'] = u'Кол-во'

        sheet['K1'] = format_date_range(prev_month_period)
        sheet['K2'] = u'Кол-во'

        sheet['L1'] = format_date(self._date)
        sheet['L2'] = u'Кол-во'

        sheet['M1'] = format_date(previous_month_day(self._date))
        sheet['M2'] = u'Кол-во'

        sheet.auto_filter.ref = 'A2:M2'
        self._format_header()

    def _add_right_vertical_borders(self):
        letters = ("C", "F", "I", "K", "M")
        map(self._add_vertical_border, letters)

    def _fill_table(self):
        sheet = self._sheet
        for row, shop in enumerate(self._get_shops(), start=3):  # Fill from 3rd row
            sheet.cell(row=row, column=1).value = shop.id
            sheet.cell(row=row, column=2).value = shop.url
            sheet.cell(row=row, column=3).value = shop.name
            sheet.cell(row=row, column=4).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=4).value = shop.curr_period_circ
            sheet.cell(row=row, column=5).value = shop.prev_period_circ
            sheet.cell(row=row, column=5).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).value = calculate_dynamic(sheet.cell(row=row, column=4),
                                                                    sheet.cell(row=row, column=5))
            sheet.cell(row=row, column=7).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=7).value = shop.curr_day_circ
            sheet.cell(row=row, column=8).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=8).value = shop.prev_day_circ
            sheet.cell(row=row, column=9).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=9).value = calculate_dynamic(sheet.cell(row=row, column=7),
                                                                    sheet.cell(row=row, column=8))
            sheet.cell(row=row, column=10).value = shop.curr_period_count
            sheet.cell(row=row, column=11).value = shop.prev_period_count
            sheet.cell(row=row, column=12).value = shop.curr_day_count
            sheet.cell(row=row, column=13).value = shop.prev_day_count

    def _get_shops(self):
        raise NotImplementedError()


class PaywaySheet(Sheet):
    def _create_table_header(self):
        sheet = self._sheet
        sheet['A1'] = u'ID ПН'
        sheet['B1'] = u'Название ПС (ID)'
        sheet['C1'] = u'Название ПН'
        sheet['D1'] = format_date_range(get_period_range(self._date))
        sheet['D2'] = u'Оборот'

        prev_month_period = get_period_range(previous_month_day(self._date))
        sheet['E1'] = format_date_range(prev_month_period)
        sheet['E2'] = u'Оборот'

        sheet['F1'] = self.dynamic_date
        sheet['F2'] = u'Динамика'

        sheet['G1'] = format_date(self._date)
        sheet['G2'] = u'Оборот'

        sheet['H1'] = format_date(previous_month_day(self._date))
        sheet['H2'] = u'Оборот'

        sheet['I1'] = self.dynamic_date
        sheet['I2'] = u'Динамика'

        sheet['J1'] = format_date_range(get_period_range(self._date))
        sheet['J2'] = u'Кол-во'

        sheet['K1'] = format_date_range(prev_month_period)
        sheet['K2'] = u'Кол-во'

        sheet['L1'] = format_date(self._date)
        sheet['L2'] = u'Кол-во'

        sheet['M1'] = format_date(previous_month_day(self._date))
        sheet['M2'] = u'Кол-во'

        sheet.auto_filter.ref = 'A2:M2'
        self._format_header()

    def _add_right_vertical_borders(self):
        letters = ("C", "F", "I", "K", "M")
        map(self._add_vertical_border, letters)

    def _fill_table(self):
        sheet = self._sheet
        for row, pw in enumerate(self._get_payways(), start=3):  # Fill from 3rd row
            sheet.cell(row=row, column=1).value = pw.id
            sheet.cell(row=row, column=2).value = '%s(%s)' % (pw.paysystem.name, pw.paysystem_id)
            sheet.cell(row=row, column=3).value = pw.name
            sheet.cell(row=row, column=4).value = pw.curr_period_circ
            sheet.cell(row=row, column=4).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=5).value = pw.prev_period_circ
            sheet.cell(row=row, column=5).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).value = calculate_dynamic(sheet.cell(row=row, column=4),
                                                                    sheet.cell(row=row, column=5))
            sheet.cell(row=row, column=6).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=7).value = pw.curr_day_circ
            sheet.cell(row=row, column=7).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=8).value = pw.prev_day_circ
            sheet.cell(row=row, column=8).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=9).value = calculate_dynamic(sheet.cell(row=row, column=7),
                                                                    sheet.cell(row=row, column=8))
            sheet.cell(row=row, column=9).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=10).value = pw.curr_period_count
            sheet.cell(row=row, column=11).value = pw.prev_period_count
            sheet.cell(row=row, column=12).value = pw.curr_day_count
            sheet.cell(row=row, column=13).value = pw.prev_day_count

    def _get_payways(self):
        raise NotImplementedError()


class PaysystemSheet(Sheet):
    def _create_table_header(self):
        sheet = self._sheet
        sheet['A1'] = u'ID ПС'
        sheet['B1'] = u'Название ПС'
        sheet['C1'] = format_date_range(get_period_range(self._date))
        sheet['C2'] = u'Оборот'

        prev_month_period = get_period_range(previous_month_day(self._date))
        sheet['D1'] = format_date_range(prev_month_period)
        sheet['D2'] = u'Оборот'

        sheet['E1'] = self.dynamic_date
        sheet['E2'] = u'Динамика'

        sheet['F1'] = format_date(self._date)
        sheet['F2'] = u'Оборот'

        sheet['G1'] = format_date(previous_month_day(self._date))
        sheet['G2'] = u'Оборот'

        sheet['H1'] = self.dynamic_date
        sheet['H2'] = u'Динамика'

        sheet['I1'] = format_date_range(get_period_range(self._date))
        sheet['I2'] = u'Кол-во'

        sheet['J1'] = format_date_range(prev_month_period)
        sheet['J2'] = u'Кол-во'

        sheet['K1'] = format_date(self._date)
        sheet['K2'] = u'Кол-во'

        sheet['L1'] = format_date(previous_month_day(self._date))
        sheet['L2'] = u'Кол-во'

        sheet.auto_filter.ref = 'A2:L2'
        self._format_header()

    def _create_table_footer(self):
        sheet = self._sheet
        max_row = sheet.max_row
        last_row = max_row + 2  # Skip two lines after last raw

        sheet['B%s' % last_row] = u'ИТОГО'
        cell_chars = ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L')
        for char in sorted(cell_chars):
            sheet['{0}{1}'.format(char, last_row)] = u'=SUM({0}3:{1}{2})'.format(char, char, max_row)

        sheet['C%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['D%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['E%s' % last_row] = calculate_dynamic(sheet['C%s' % last_row], sheet['D%s' % last_row])
        sheet['E%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT
        sheet['F%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['G%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['H%s' % last_row] = calculate_dynamic(sheet['F%s' % last_row], sheet['G%s' % last_row])
        sheet['H%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT

        self._apply_border()

    def _add_right_vertical_borders(self):
        letters = ("B", "E", "H", "J", "L")
        map(self._add_vertical_border, letters)

    def _fill_table(self):
        sheet = self._sheet
        for row, ps in enumerate(self._get_paysystems(), start=3):  # Fill from 3rd row
            sheet.cell(row=row, column=1).value = ps.id
            sheet.cell(row=row, column=2).value = ps.name
            sheet.cell(row=row, column=3).value = ps.curr_period_circ
            sheet.cell(row=row, column=3).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=4).value = ps.prev_period_circ
            sheet.cell(row=row, column=4).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=5).value = calculate_dynamic(sheet.cell(row=row, column=3),
                                                                    sheet.cell(row=row, column=4))
            sheet.cell(row=row, column=5).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).value = ps.curr_day_circ
            sheet.cell(row=row, column=6).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=7).value = ps.prev_day_circ
            sheet.cell(row=row, column=7).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=8).value = calculate_dynamic(sheet.cell(row=row, column=6),
                                                                    sheet.cell(row=row, column=7))
            sheet.cell(row=row, column=8).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=9).value = ps.curr_period_count
            sheet.cell(row=row, column=10).value = ps.prev_period_count
            sheet.cell(row=row, column=11).value = ps.curr_day_count
            sheet.cell(row=row, column=12).value = ps.prev_day_count

    def _get_paysystems(self):
        raise NotImplementedError()


class PaymethodSheet(Sheet):
    def _create_table_header(self):
        sheet = self._sheet
        sheet['A1'] = u'ID ПМ'
        sheet['B1'] = u'Название метода оплаты'
        sheet['C1'] = format_date_range(get_period_range(self._date))
        sheet['C2'] = u'Оборот'

        prev_month_period = get_period_range(previous_month_day(self._date))
        sheet['D1'] = format_date_range(prev_month_period)
        sheet['D2'] = u'Оборот'

        sheet['E1'] = self.dynamic_date
        sheet['E2'] = u'Динамика'

        sheet['F1'] = format_date(self._date)
        sheet['F2'] = u'Оборот'

        sheet['G1'] = format_date(previous_month_day(self._date))
        sheet['G2'] = u'Оборот'

        sheet['H1'] = self.dynamic_date
        sheet['H2'] = u'Динамика'

        sheet['I1'] = format_date_range(get_period_range(self._date))
        sheet['I2'] = u'Кол-во'

        sheet['J1'] = format_date_range(prev_month_period)
        sheet['J2'] = u'Кол-во'

        sheet['K1'] = format_date(self._date)
        sheet['K2'] = u'Кол-во'

        sheet['L1'] = format_date(previous_month_day(self._date))
        sheet['L2'] = u'Кол-во'

        sheet.auto_filter.ref = 'A2:L2'
        self._format_header()

    def _create_table_footer(self):
        sheet = self._sheet
        max_row = sheet.max_row
        last_row = max_row + 2  # Skip two lines after last raw

        sheet['B%s' % last_row] = u'ИТОГО'
        cell_chars = ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L')
        for char in sorted(cell_chars):
            sheet['{0}{1}'.format(char, last_row)] = u'=SUM({0}3:{1}{2})'.format(char, char, max_row)

        sheet['C%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['D%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['E%s' % last_row] = calculate_dynamic(sheet['C%s' % last_row], sheet['D%s' % last_row])
        sheet['E%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT
        sheet['F%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['G%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['H%s' % last_row] = calculate_dynamic(sheet['F%s' % last_row], sheet['G%s' % last_row])
        sheet['H%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT

        self._apply_border()

    def _add_right_vertical_borders(self):
        letters = ("B", "E", "H", "J", "L")
        map(self._add_vertical_border, letters)

    def _fill_table(self):
        sheet = self._sheet
        for row, pm in enumerate(self._get_paymethods(), start=3):  # Fill from 3rd row
            sheet.cell(row=row, column=1).value = pm.id
            sheet.cell(row=row, column=2).value = pm.name
            sheet.cell(row=row, column=3).value = pm.curr_period_circ
            sheet.cell(row=row, column=3).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=4).value = pm.prev_period_circ
            sheet.cell(row=row, column=4).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=5).value = calculate_dynamic(sheet.cell(row=row, column=3),
                                                                    sheet.cell(row=row, column=4))
            sheet.cell(row=row, column=5).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).value = pm.curr_day_circ
            sheet.cell(row=row, column=6).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=7).value = pm.prev_day_circ
            sheet.cell(row=row, column=7).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=8).value = calculate_dynamic(sheet.cell(row=row, column=6),
                                                                    sheet.cell(row=row, column=7))
            sheet.cell(row=row, column=8).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=9).value = pm.curr_period_count
            sheet.cell(row=row, column=10).value = pm.prev_period_count
            sheet.cell(row=row, column=11).value = pm.curr_day_count
            sheet.cell(row=row, column=12).value = pm.prev_day_count

    def _get_paymethods(self):
        raise NotImplementedError()


class ProjectSheet(Sheet):
    def _create_table_header(self):
        sheet = self._sheet
        sheet['A1'] = u'ID проекта'
        sheet['B1'] = u'URL проекта'
        sheet['C1'] = format_date_range(get_period_range(self._date))
        sheet['C2'] = u'Оборот'

        prev_month_period = get_period_range(previous_month_day(self._date))
        sheet['D1'] = format_date_range(prev_month_period)
        sheet['D2'] = u'Оборот'

        sheet['E1'] = self.dynamic_date
        sheet['E2'] = u'Динамика'

        sheet['F1'] = format_date(self._date)
        sheet['F2'] = u'Оборот'

        sheet['G1'] = format_date(previous_month_day(self._date))
        sheet['G2'] = u'Оборот'

        sheet['H1'] = self.dynamic_date
        sheet['H2'] = u'Динамика'

        sheet['I1'] = format_date_range(get_period_range(self._date))
        sheet['I2'] = u'Кол-во'

        sheet['J1'] = format_date_range(prev_month_period)
        sheet['J2'] = u'Кол-во'

        sheet['K1'] = format_date(self._date)
        sheet['K2'] = u'Кол-во'

        sheet['L1'] = format_date(previous_month_day(self._date))
        sheet['L2'] = u'Кол-во'

        sheet.auto_filter.ref = 'A2:L2'
        self._format_header()

    def _create_table_footer(self):
        sheet = self._sheet
        max_row = sheet.max_row
        last_row = max_row + 2  # Skip two lines after last raw

        sheet['B%s' % last_row] = u'ИТОГО'
        cell_chars = ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L')
        for char in sorted(cell_chars):
            sheet['{0}{1}'.format(char, last_row)] = u'=SUM({0}3:{1}{2})'.format(char, char, max_row)

        sheet['C%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['D%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['E%s' % last_row] = calculate_dynamic(sheet['C%s' % last_row], sheet['D%s' % last_row])
        sheet['E%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT
        sheet['F%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['G%s' % last_row].number_format = CIRC_NUMBER_FORMAT
        sheet['H%s' % last_row] = calculate_dynamic(sheet['F%s' % last_row], sheet['G%s' % last_row])
        sheet['H%s' % last_row].number_format = DYNAMIC_NUMBER_FORMAT

        self._apply_border()

    def _add_right_vertical_borders(self):
        letters = ("B", "E", "H", "J", "L")
        map(self._add_vertical_border, letters)

    def _fill_table(self):
        sheet = self._sheet
        for row, p in enumerate(self._get_projects(), start=3):  # Fill from 3rd row
            sheet.cell(row=row, column=1).value = p.id
            sheet.cell(row=row, column=2).value = p.url
            sheet.cell(row=row, column=3).value = p.curr_period_circ
            sheet.cell(row=row, column=3).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=4).value = p.prev_period_circ
            sheet.cell(row=row, column=4).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=5).value = calculate_dynamic(sheet.cell(row=row, column=3),
                                                                    sheet.cell(row=row, column=4))
            sheet.cell(row=row, column=5).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=6).value = p.curr_day_circ
            sheet.cell(row=row, column=6).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=7).value = p.prev_day_circ
            sheet.cell(row=row, column=7).number_format = CIRC_NUMBER_FORMAT
            sheet.cell(row=row, column=8).value = calculate_dynamic(sheet.cell(row=row, column=6),
                                                                    sheet.cell(row=row, column=7))
            sheet.cell(row=row, column=8).number_format = DYNAMIC_NUMBER_FORMAT
            sheet.cell(row=row, column=9).value = p.curr_period_count
            sheet.cell(row=row, column=10).value = p.prev_period_count
            sheet.cell(row=row, column=11).value = p.curr_day_count
            sheet.cell(row=row, column=12).value = p.prev_day_count

    def _get_projects(self):
        raise NotImplementedError()
