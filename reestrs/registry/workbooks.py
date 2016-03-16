# -*- coding: utf-8 -*-

import logging
from collections import OrderedDict

import openpyxl

from models import Currency, Shop
from utils import format_date
from .sheets import Sheet
from config import INVOICE_STATISTIC_DIR, WITHDRAW_STATISTIC_DIR
from .invoice_sheets import ShopInvoiceSheet, PaywayInvoiceSheet, PaysystemInvoiceSheet, PaymethodInvoiceSheet,\
    ProjectInvoiceSheet
from .withdraw_sheets import ShopWithdrawSheet, PaywayWithdrawSheet, PaysystemWithdrawSheet, PaymethodWithdrawSheet


log = logging.getLogger(__name__)


def verify_shop(shop_id):
    shop = Shop.get_by_id(shop_id)
    if shop is None:
        raise ValueError('Магазин с shop_id=%s не существует' % shop_id)


class WorkBook(object):
    def __init__(self, name, date):
        self._date = date
        self.name = '%s %s.xlsx' % (name, format_date(self._date))
        self._wb = openpyxl.Workbook()

        # Remove dummy sheet
        if self._wb.worksheets:
            self._wb.remove_sheet(self._wb.worksheets[0])

        self.create_sheets()
        self.save()

    def create_sheet(self, title, currency):
        return Sheet(self._wb, title, currency, self._date)

    def create_sheets(self):
        for code, alias in self._get_currencies().iteritems():
            title = '{0}({1})'.format(alias.upper(), code)
            self.create_sheet(title, code)

    def save(self):
        raise NotImplementedError()

    def _get_currencies(self):
        currencies = {currency.code: currency.alias for currency in Currency.select()}
        return OrderedDict(sorted(currencies.items()))


class InvoiceWorkbook(WorkBook):
    def save(self):
        self.location = INVOICE_STATISTIC_DIR + self.name
        self._wb.save(self.location)
        log.info("Файл %s успешно создан" % self.location)


class WithdrawWorkbook(WorkBook):
    def save(self):
        self.location = WITHDRAW_STATISTIC_DIR + self.name
        self._wb.save(self.location)
        log.info("Файл %s успешно создан" % self.location)


class ShopInvoiceWorkBook(InvoiceWorkbook):
    def create_sheet(self, title, currency):
        return ShopInvoiceSheet(self._wb, title, currency, self._date)


class PaywayInvoiceWorkBook(InvoiceWorkbook):
    def create_sheet(self, title, currency):
        return PaywayInvoiceSheet(self._wb, title, currency, self._date)


class PaysytemInvoiceWorkBook(InvoiceWorkbook):
    def create_sheet(self, title, currency):
        return PaysystemInvoiceSheet(self._wb, title, currency, self._date)


class PaymethodInvoiceWorkBook(InvoiceWorkbook):
    def create_sheet(self, title, currency):
        return PaymethodInvoiceSheet(self._wb, title, currency, self._date)


class ShopWithdrawWorkBook(WithdrawWorkbook):
    def create_sheet(self, title, currency):
        return ShopWithdrawSheet(self._wb, title, currency, self._date)


class PaywayWithdrawWorkBook(WithdrawWorkbook):
    def create_sheet(self, title, currency):
        return PaywayWithdrawSheet(self._wb, title, currency, self._date)


class PaysytemWithdrawWorkBook(WithdrawWorkbook):
    def create_sheet(self, title, currency):
        return PaysystemWithdrawSheet(self._wb, title, currency, self._date)


class PaymethodWithdrawWorkBook(WithdrawWorkbook):
    def create_sheet(self, title, currency):
        return PaymethodWithdrawSheet(self._wb, title, currency, self._date)


class ProjectInvoiceWorkBook(InvoiceWorkbook):
    def __init__(self, name, date, shop_id):
        verify_shop(shop_id)
        self._shop_id = shop_id
        super(ProjectInvoiceWorkBook, self).__init__(name, date)

    def create_sheet(self, title, currency):
        return ProjectInvoiceSheet(self._wb, title, currency, self._date, self._shop_id)
